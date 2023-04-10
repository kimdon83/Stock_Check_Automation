# %%
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from pandas._libs.tslibs import NaT
from pandas.core.arrays.sparse import dtype
from sqlalchemy import create_engine
from sqlalchemy.engine import URL
from datetime import datetime
from dateutil.relativedelta import *
import time
import matplotlib.pyplot as plt
import matplotlib as mpl

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# %%
timelist = []
# %%
# Connect to KIRA server
start = time.time()

import json

with open(r'C:\Users\KISS Admin\Desktop\IVYENT_DH\data.json', 'r') as f:
    data = json.load(f)

# getID and password
server = data['server']
database = data['database']
username = data['username']
password = data['password']
connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + \
    server+';DATABASE='+database+';UID='+username+';PWD=' + password
connection_url = URL.create(
    "mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url, fast_executemany=True)
print("Connection Established:")

end = time.time()
timelist.append([end-start, "Connect to KIRA server"])



# %%

todays = datetime.today()
# first_days = todays.replace(day=1)
# last_days = datetime(todays.year, todays.month, 1) + relativedelta(months=1) + relativedelta(seconds=-1)
# days_left = last_days - todays
today = todays.strftime('%Y-%m-%d')
curYM = todays.strftime('%Y%m')
# first_day = first_days.strftime('%Y-%m-%d')
# last_day = last_days.strftime('%Y-%m-%d')
# business_days = np.busday_count(begindates=first_day, enddates=today) #By today
# business_days_thismonth = np.busday_count(begindates=first_day, enddates=last_day)
# business_days_left = np.busday_count(begindates=today, enddates=last_day)
# %%
# read simulator_input.csv
# simulator
loc_input = r'C:\Users\KISS Admin\Desktop\stock check practice\simulator_input.csv'
input=pd.read_csv(loc_input)

#
orderlimit_df = pd.read_sql("""SELECT material, from_date, to_date FROM [ivy.mm.dim.orderlimit] WHERE from_date<=GETDATE() and to_date>=GETDATE()""", con=engine)
orderlimit_df.columns = ['material', 'from_date', 'to_date']
orderlimit_df['orderlimit'] = 1
orderlimit_df = orderlimit_df.drop_duplicates(subset='material')

# %% BOM (dimbom_aset)
bom_df = pd.read_sql("""select bom_parent_material as material from [ivy.mm.dim.bom_aset] GROUP BY bom_parent_material""", con=engine)
bom_df['bom'] = 1 

# %% dim.mtrl for ms
mtrl_df = pd.read_sql("""select material, ms from [ivy.mm.dim.mtrl] """, con=engine)

# %% Final_df : master table
merge4_df = pd.merge(input, orderlimit_df, on='material', how='left') #if order limit, then orderlimit column == 1
merge5_df = pd.merge(merge4_df, mtrl_df, on='material', how='left') #add ms
merge6_df = pd.merge(merge5_df, bom_df, on='material', how='left') #if bom, then bom column == 1

# %%
final_df = merge6_df[['material', 'plant', 'qty','orderlimit', 'bom','ms']]
final_df.columns = ['material', 'plant', 'qty','orderlimit', 'bom','ms']
final_df.insert(5,'availability','None')
final_df.loc[:,['orderlimit','bom']] = final_df[['orderlimit','bom']].fillna(0)
final_df.loc[:,'orderlimit'] = final_df['orderlimit'].astype('int')
final_df.loc[:,'bom'] = final_df['bom'].astype('int')
final_df.loc[:,'plant'] = final_df['plant'].astype('str')
final_df.reset_index(inplace=True)
final_df.drop(['index'], axis=1, inplace=True)

# %%
#Ivy

salesorg= str(input.loc[0,"salesorg"])
if salesorg == '1300': #For AST orders, we do not check for plant 1000
    plant_list = ['1100', '1110']

else:
    plant_list = ['1000', '1100', '1110']

final_df = final_df[final_df['plant'].isin(plant_list)]

if "order_number" in input.columns:
    order_number=input.loc[0,'order_number']
    final_df.insert(7,'order_number',order_number)
print(final_df)

final_df.loc[:,'plant'] = final_df['plant'].astype('int')

input=final_df.copy()
#
inputKDC=input.loc[input.plant%100==0]
inputLA=input.loc[input.plant%100!=0]
input_order =input.reset_index()

# %%
# %%
df_mtrl= pd.read_sql("""SELECT material, ms, pdt FROM [ivy.mm.dim.mtrl]""", con=engine)
df_mtrl.head()

df_po = pd.read_sql("""SELECT material, act_date, sum(po_qty+asn_qty) as poasn_qty FROM [ivy.mm.dim.fact_poasn]
GROUP BY material, act_date
""", con=engine)
df_po.head()


# %%
# get the full table for this calcutation.
################################################################

if(len(inputKDC)>0):
    targetPlant='simulate_KDC'

if(len(inputLA)>0):
    targetPlant='simulate_LA'

def simulate_KDC_LA(targetPlant):
    start = time.time()
    print("get sql date for",targetPlant)
    sql_string="""
    DECLARE @mthwitdh AS INT
    DECLARE @3Mwds AS FLOAT

    SELECT @mthwitdh = 7;

    SELECT @3Mwds = (
            SELECT COUNT(*) AS WDs
            FROM [ivy.mm.dim.date]
            WHERE IsKissHoliday != 1 AND thedate BETWEEN DATEADD(MM, - 3, DATEADD(DD, - 1, GETDATE())) AND DATEADD(DD, - 1, GETDATE())
            GROUP BY IsKissHoliday
            );

    WITH pppDailyThisMonth
    AS (
        SELECT SUM(qty) AS thisMthReOdqty, material, plant
        FROM [ivy.sd.fact.bill_ppp]
        WHERE act_date BETWEEN DATEADD(DD, 1, EOMONTH(GETDATE(), - 1)) AND GETDATE() AND ordsqc > 1   
        and material in ('change_string')
        GROUP BY material, plant
        ), ppp
        --avgMreorder within 3month, material, plant FROM [ivy.sd.fact.bill_ppp]
    AS (
        SELECT SUM(qty) AS reorder3M, material, plant
        FROM [ivy.sd.fact.bill_ppp]
        WHERE act_date BETWEEN DATEADD(MM, - 3, DATEADD(DD, - 1, GETDATE())) AND DATEADD(DD, - 1, GETDATE()) AND ordsqc > 1
        and material in ('change_string')
        GROUP BY material, plant
        ), backOrder
        -- avgMbo within 3month, material, plant FROM [ivy.sd.fact.bo] 
    AS (
        SELECT SUM(bo_qty) AS bo3M, material, plant
        FROM [ivy.sd.fact.bo]
        WHERE (act_date BETWEEN DATEADD(MM, - 3, DATEADD(DD, - 1, GETDATE())) AND DATEADD(DD, - 1, GETDATE()))
        and material in ('change_string')
        GROUP BY material, plant
        ), pppbo
    AS (
        SELECT cast(reorder3M AS FLOAT) / @3Mwds AS reorderPerWDs, T1.material, T1.plant, cast(bo3M AS FLOAT) / @3Mwds AS boPerWDs
        FROM ppp T1
        LEFT JOIN backOrder T2 ON T1.material = T2.material AND T1.plant = T2.plant
            --ORDER BY plant, material
        ), T4fcst
        -- Table to make fcst table. FROM this month to upcoming 5 monthl
    AS (
        SELECT material, SUM(eship) AS eship, FORMAT(act_date, 'MMyyyy') AS MMYYYY, plant
        FROM [ivy.mm.dim.factfcst]
        WHERE act_date BETWEEN DATEADD(DD, - DAY(GETDATE()), GETDATE()) AND DATEADD(MM, @mthwitdh + 1, DATEADD(DD, - DAY(GETDATE()), GETDATE()))
        and material in ('change_string')                            
        GROUP BY material, FORMAT(act_date, 'MMyyyy'), plant
        ), fcst
    AS (
        SELECT T1.TheDate, T1.accumWDs, T1.MMYYYY, T1.IsKissHoliday, (1 - T3.IsKissHoliday) * (CONVERT(FLOAT, T2.eship) / T3.workdaysInMonth) AS fcstPerWDs, T2.plant, T2.material
        FROM (
            SELECT TheDate, workdaysInMonth AS WDs, workdaysInMonth - workdaysLeftInMonth AS accumWDs, MMYYYY, IsKissHoliday
            FROM [ivy.mm.dim.date]
            WHERE thedate BETWEEN DATEADD(DD, - DAY(GETDATE()), GETDATE()) AND DATEADD(MM, @mthwitdh + 1, DATEADD(DD, - DAY(GETDATE()), GETDATE()))
            ) T1
        LEFT JOIN T4fcst T2 ON T1.MMYYYY = T2.MMYYYY
        LEFT JOIN [ivy.mm.dim.date] T3 on T1.TheDate=T3.TheDate
        WHERE T1.thedate BETWEEN DATEADD(DAY, - 6, GETDATE()) AND DATEADD(MONTH, @mthwitdh, GETDATE())
        ), Tpoasn
    AS (
        SELECT material, plant, act_date, sum(po_qty) AS po_qty, sum(asn_qty) AS asn_qty
        FROM [ivy.mm.dim.fact_poasn]
        -- WHERE po_num NOT LIKE '43%' -- exclude intra_company po not exclude for individual plant
        WHERE material in ('change_string')
        GROUP BY material, plant, act_date
        ), mrp01
    AS (
        SELECT *
        FROM [ivy.mm.dim.mrp01]
        WHERE pgr != 'IEC' -- exclude IEC for total stock
        and material in ('change_string')
        ), TOTAL
    AS (
        SELECT T2.PL_plant, T1.thedate, T3.material, T3.nsp, T1.IsKissHoliday, T6.boPerWDs, T5.po_qty + T5.asn_qty AS poasn_qty, 
        T6.reorderPerWDs, T8.total_stock - T8.blocked - T8.subcont_qty AS On_hand_qty, T9.fcstPerWDs, 
        T9.accumWDs, T10.thisMthReOdqty
        FROM (
            SELECT DISTINCT PL_PLANT -- pl_plant
            FROM [ivy.mm.dim.mrp01]
            ) T2
        CROSS JOIN (
            SELECT THEDATE, IsKissHoliday
            FROM [ivy.mm.dim.date]
            WHERE thedate BETWEEN DATEADD(DAY, - 6, GETDATE()) AND DATEADD(MONTH, @mthwitdh, GETDATE())
            ) T1
        CROSS JOIN (
            SELECT MATERIAL, nsp --material
            FROM [ivy.mm.dim.mtrl]
            WHERE material in ('change_string')
            ) T3
        LEFT JOIN Tpoasn T5 ON T3.material = T5.material -- poasn_qty
            AND T2.pl_plant = T5.plant AND T1.TheDate = T5.act_date
        LEFT JOIN pppbo T6 ON T3.material = T6.material -- average Monthly reorder qty
            AND T2.pl_plant = T6.plant
        LEFT JOIN [ivy.mm.dim.mrp01] T8 ON T3.material = T8.material -- on_hand qty
            AND T2.pl_plant = T8.pl_plant
        LEFT JOIN fcst T9 ON T3.material = T9.material -- fcstPerWDs, IsKissHoliday
            AND T2.pl_plant = T9.plant AND T9.TheDate = T1.TheDate
        LEFT JOIN pppDailyThisMonth T10 ON T10.material = T3.material AND T10.plant = T2.pl_plant
            -- WHERE T8.pgr != 'IEC' -- exclude IEC for total stock
            --LEFT JOIN [ivy.sd.fact.order] od ON od.act_date= T1.TheDate and od.material=T3.material
        ), TOTAL2
        -- NULL value to 0 (avgDbo, poasn_qty, avgDreorder, fcstPerWDs,On_hand_qty)
    AS (
        SELECT pl_plant, TheDate, material, 
        CASE WHEN nsp IS NULL THEN 0 ELSE nsp END AS nsp, 
        CASE WHEN (boPerWDs IS NULL) THEN 0 ELSE boPerWDs END AS avgDbo, 
        CASE WHEN (poasn_qty IS NULL) THEN 0 ELSE poasn_qty END AS poasn_qty, 
        CASE WHEN (reorderPerWDs IS NULL) THEN 0 ELSE reorderPerWDs END AS avgDreorder, 
        CASE WHEN (fcstPerWDs IS NULL) THEN 0 ELSE fcstPerWDs END AS fcstPerWDs, 
        CASE WHEN (On_hand_qty IS NULL) THEN 0 ELSE On_hand_qty END AS On_hand_qty, 
        CASE WHEN (thisMthReOdqty IS NULL) THEN 0 ELSE thisMthReOdqty END AS thisMthReOdqty, 
            IsKissHoliday, accumWDs
        FROM Total
        )
    SELECT pl_plant AS plant, TheDate, material AS mtrl, nsp, avgDbo, poasn_qty, avgDreorder, On_hand_qty, 
        CASE WHEN (fcstPerWDs = 0 AND IsKissHoliday = 0 AND pl_plant IN ('1100', '1400','G140','G110')) THEN 
                    avgDreorder + avgDbo ELSE fcstPerWDs END AS fcstD, thisMthReOdqty
    FROM TOTAL2
    ORDER BY plant, mtrl, TheDate
    """
    # simulator
    if targetPlant=='simulate_KDC':        
        replace_material='\''+'\',\''.join(map(str,list(inputKDC.material)))+'\''
    elif targetPlant=='simulate_LA':
        replace_material='\''+'\',\''.join(map(str,list(inputLA.material)))+'\''    
    sql_string=sql_string.replace("'change_string'",replace_material)

    # %% 
    df_ft = pd.read_sql(sql_string, con=engine)

    print("full table is ready")

    df_wds = pd.read_sql("""
    DECLARE @mthwitdh AS INT
    SELECT @mthwitdh = 7;
    With T1 AS(
    SELECT TheDate, SUM(1 - IsWeekend) OVER (PARTITION BY MMYYYY) AS WDs, SUM(1 - IsWeekend) OVER (
            PARTITION BY MMYYYY ORDER BY TheDate
            ) AS accumWDs, IsWeekend
    FROM [ivy.mm.dim.date]
    WHERE thedate BETWEEN DATEADD(DD, - DAY(GETDATE()), GETDATE()) AND DATEADD(MM, @mthwitdh+1, DATEADD(DD, - DAY(GETDATE()), 
                        GETDATE()))
    )
    SELECT * FROM T1
    WHERE thedate BETWEEN DATEADD(DAY, - 6, GETDATE()) AND DATEADD(MONTH, @mthwitdh, GETDATE())
    ORDER BY TheDate
    """, con=engine)

    end = time.time()
    timelist.append([end-start, "Get full table from SQL server"])


    # %%
    # define location
    start = time.time()
    file_loc = r'C:\Users\KISS Admin\Desktop\IVYENT_DH\P6. DailyDM except codes\simulation'
    # group by mtrl & TheDate
    if targetPlant=='simulate_KDC':
        df_total = df_ft[df_ft.plant.isin(['1000','1100','1300','1400','G140','G110','G100','G130'])].groupby(["mtrl", "TheDate"]).agg({'nsp':'mean','avgDbo':'sum',"poasn_qty":'sum','avgDreorder':'sum','On_hand_qty':'sum','fcstD':'sum','thisMthReOdqty':'sum'})
    elif targetPlant=='simulate_LA':
        df_total = df_ft[df_ft.plant.isin(['1110','1410'])].groupby(["mtrl", "TheDate"]).agg({'nsp':'mean','avgDbo':'sum',"poasn_qty":'sum','avgDreorder':'sum','On_hand_qty':'sum','fcstD':'sum','thisMthReOdqty':'sum'})
    df_total = df_total.reset_index()
    df_total = df_total.merge(df_wds, how='left', on='TheDate')
    # left join with df_wds

    # reduce on_hand_qty as input requirement qty. 
    # simulator
    for index, row in df_total.iterrows():
        if targetPlant=='simulate_KDC':
            row.On_hand_qty=row.On_hand_qty-inputKDC[inputKDC['material']==row.mtrl].qty
        elif targetPlant=='simulate_LA':
            row.On_hand_qty=row.On_hand_qty-inputLA[inputKDC['material']==row.mtrl].qty       

    end = time.time()
    timelist.append(
        [end-start, """df_total = df_ft.groupby(["mtrl", "TheDate"]).sum()"""])

    # %%
    # define DailyCalculate
    start = time.time()

    def DailyCalculate(df):
        half_flag = False
        if int(todays.strftime('%d')) > 15:
            print("The adjustment after half month on the starting month will be applied")
            half_flag = True
        # df = df.reset_index()

        df_mtrl = pd.DataFrame(df["mtrl"].unique())
        df_date = pd.DataFrame(df["TheDate"].unique())

        # set BOseq, residue, BOqty on df
        df["BOseq"] = 999
        df["residue"] = 999
        df["BOqty"] = 0
        df["BO$"] = 0
        df = df[['mtrl', 'TheDate', 'nsp', 'avgDbo', 'poasn_qty', 'avgDreorder', 'On_hand_qty', 'fcstD',
                "BOseq", "residue", "BOqty", "BO$", 'thisMthReOdqty', 'WDs', 'accumWDs']]
        # define po processing time as 5days
        poDays = 5

        df_mtrl = df_mtrl.to_numpy()

        colnames = df.columns
        df = df.to_numpy()

        for index_mtrl in range(len(df_mtrl)):
            if index_mtrl % 19 == 0:
                # print( f'{df_mtrl.loc[index_mtrl][0]:15} {float(index_mtrl+1)/float(len(df_mtrl))*100:.2f}% ') # print % progress
                # print % progress
                print(
                    f'{df_mtrl[index_mtrl][0]:15} {float(index_mtrl+1)/float(len(df_mtrl))*100:.2f}% ')
            # set current BOflag, BOseq, Residue
            BOflag = 0
            curBOseq = 0
            # df.loc[index_mtrl*len(df_date), "On_hand_qty"]
            curResidue = df[index_mtrl*len(df_date)][6]
            # check if there is no poasn for this mtrl
            # poasn_test = df.loc[df["mtrl"] == df.loc[index_mtrl *len(df_date), "mtrl"], "poasn_qty"].sum() == 0
            df[index_mtrl*len(df_date):(index_mtrl+1) * len(df_date), 4].sum() ==0

            poasn_test = df[index_mtrl*len(df_date):(index_mtrl+1) * len(df_date), 4].sum() ==0
            if ((curResidue == 0) & poasn_test):  # if no inventory and poasn => set residue:0 and BOseq:-1
                # df.loc[index_mtrl*len(df_date):(index_mtrl+1) *	len(df_date)-1, "residue"] = 0
                df[index_mtrl*len(df_date):(index_mtrl+1) * len(df_date), 9] = 0
                # df.loc[index_mtrl*len(df_date):(index_mtrl+1) * len(df_date)-1, "BOseq"] = -1
                df[index_mtrl*len(df_date):(index_mtrl+1) * len(df_date), 8] = -1
            else:
                # df[index_mtrl*len(df_date):(index_mtrl+1)*len(df_date)]=CalcMtrl(df,index_mtrl,len(df_date),poDays,curBOseq,BOflag,curResidue)
                for index_date in range(5, len(df_date)):

                    curIndex = index_mtrl*len(df_date)+index_date  # current Index
                    # df.loc[curIndex,"On_hand_qty"]=curResidue+df.loc[curIndex-poDays,"poasn_qty"]
                    curYMflag = (curYM == df[curIndex, 1].strftime('%Y%m'))
                    if half_flag == True:  # halfday
                        # if df[curIndex,"fcstD"]*df[curIndex,accumWDs]<df[curIndex,thisMthReOdqty]:
                        if (df[curIndex, 7]*df[curIndex, 14] < df[curIndex, 12]) & curYMflag:
                            # print(f'{df[curIndex,7]}*{df[curIndex,14]}<{df[curIndex,12]}')
                            df[curIndex, 7] = df[curIndex, 12]/df[curIndex, 14]
                    df[curIndex, 6] = curResidue+df[curIndex-poDays, 4]
                    if BOflag == 1:  # BO status
                        # poasn comes => end of BO, set BOflag, BOseq as out of BO. calc. curResidue
                        if df[curIndex-poDays, 4] > 0:
                            BOflag = 0
                            # df.loc[curIndex, "BOseq"] = 0
                            df[curIndex, 8] = 0
                            # curResidue = curResidue + df.loc[curIndex-poDays, "poasn_qty"]-df.loc[curIndex, "fcstD"]
                            curResidue = curResidue + \
                                df[curIndex-poDays, 4]-df[curIndex, 7]
                        else:
                            df[curIndex, 8] = curBOseq
                    else:  # not BO
                        # curResidue = curResidue + df.loc[curIndex-poDays, "poasn_qty"]-df.loc[curIndex,"fcstD"]
                        curResidue = curResidue + \
                            df[curIndex-poDays, 4]-df[curIndex, 7]
                        # Start of BO. +=1 BOseq. set curResidue, BOflag according to BO.
                        if curResidue <= 0:
                            curBOseq += 1
                            curResidue = 0
                            BOflag = 1
                            df[curIndex, 8] = curBOseq
                        else:  # curResidue >0 -> not BO
                            df[curIndex, 8] = 0
                    # df.loc[curIndex, "residue"] = curResidue
                    df[curIndex, 9] = curResidue

                    # For BO days, set BOqty as fcstD, calc. BO$ = BOqty * nsp
                    if df[curIndex, 8] != 0:
                        # df.loc[curIndex, "BOqty"]=df.loc[curIndex, "fcstD"]
                        df[curIndex, 10] = df[curIndex, 7]
                        # df.loc[curIndex, "BO$"]=df.loc[curIndex, "BOqty"]*df.loc[curIndex, "nsp"]
                        df[curIndex, 11] = df[curIndex, 10]*df[curIndex, 2]
                    else:
                        df[curIndex, 10] = 0
        print("creating The result table was done")

        # calculate BO$. save Total DM table
        df = pd.DataFrame(df)
        df.columns = colnames

        df["BO$"] = df["BOqty"]*df["nsp"]
        # df=df.loc[df.BOseq!=999]
        return df


    end = time.time()
    timelist.append([end-start, "def DailyCalculate(df):"])

    # %%
    # apply DailyCalculate to df_total and save as csv

    start = time.time()
    todays = datetime.today().date()
    today = todays.strftime('%Y-%m-%d')

    df_total = DailyCalculate(df_total)
    df_total = df_total[df_total['TheDate'] >= todays]

    df_total.reset_index(inplace=True)
    df_total.drop("index", axis=1, inplace=True)

    total_loc = file_loc+"\\"+today+"_"+targetPlant+"_ESA.csv"

    # select to columns for export
    df_total = df_total[['mtrl', 'TheDate', 'On_hand_qty',
                        'residue', 'fcstD',  'BOqty', 'BO$', 'BOseq']].copy()
    df_total[['On_hand_qty','residue', 'fcstD',  'BOqty', 'BO$']]=df_total[['On_hand_qty','residue', 'fcstD',  'BOqty', 'BO$']].astype(np.float64).round(2)

    df_total['loc']='simulation'
    df_total.to_csv(total_loc, index=False)
    print('exporting ESA.csv was done')

    end = time.time()
    timelist.append([end-start, "caluculate Daily and to_csv result"])

    # %%
    df_mtrl= pd.read_sql("""SELECT material, ms, pdt FROM [ivy.mm.dim.mtrl]""", con=engine)
    df_mtrl.head()

    df_po = pd.read_sql("""SELECT material, act_date, sum(po_qty+asn_qty) as poasn_qty FROM [ivy.mm.dim.fact_poasn]
    WHERE plant in ('1100','1400','G110','G140','1000','G100')
    GROUP BY material, act_date
    """, con=engine)
    df_po.head()

    # %%
    # group by mtrl and BOseq to show summary data of BOdates and BOqty,BO$
    # plot the BOdates, save the summary csv and png file
    start = time.time()
    todays = datetime.today()
    today = todays.strftime('%Y-%m-%d')
    # total_loc = file_loc+"\\"+today+"_"+targetPlant+"_ESA.csv"
    df_total = pd.read_csv(total_loc)

    df_result = df_total.groupby(['mtrl', 'BOseq']).agg(
        {'TheDate': ['min', 'count', 'max'], 'BOqty': ['sum'], 'BO$': 'sum'})
    df_result = df_result.reset_index()
    df_result.columns = ['mtrl', 'BOseq', 'StartDate',
                        '#ofBOdays', 'EndDate', 'BOqty', 'BO$']

    # df_result1 = df_result[df_result.BOseq != 0].copy()
    df_result1 = df_result.copy()
    df_result1.loc[:, "StartDate"] = df_result1.loc[:,
                                                    "StartDate"].apply(lambda x: pd.to_datetime(x))
    df_result1["EndDate"] = df_result1.loc[:, "EndDate"].apply(
        lambda x: pd.to_datetime(x))

    # df_result = summary_DM(df_total)
    df_result1 = df_result1[['mtrl', 'BOseq', 'StartDate', 'EndDate',
                            '#ofBOdays', 'BOqty', 'BO$']]

    if len(df_result1[df_result.BOseq != 0]) > 0:
        df_result1 = df_result1[df_result.BOseq != 0].copy()
        result_loc = file_loc+"\\"+today+"_"+targetPlant+"_BO.csv"

        # add ms , pdt to df_result1 from df_mtrl
        df_result1=df_result1.merge(df_mtrl, how='left', left_on='mtrl',right_on='material')
        df_result1.drop("material", axis=1, inplace=True)

        df_result1['pdt']=df_result1.apply(lambda row: 90 if \
            (row.pdt<90)&(row.ms in {'01','91','41'}) else row.pdt, axis=1)

        df_result1['bo_bf_pdt'] =df_result1.apply(lambda row: "yes" if (todays.date() \
            + timedelta(days=row['pdt'])) > row['StartDate'].date() else "no", axis=1)

        row= df_result1.loc[1] # for degub  

        df_result1["po_date"]=''
        df_result1["poasn_qty"]=''

        # df_result1=df_result1.merge(df_first_po, how='left', left_on='mtrl',right_on='material').drop('material',axis=1)
        for index,row in df_result1.iterrows():
            po_next_bo =df_po[(row.StartDate.date()<df_po.loc[:,"act_date"] ) & (row.mtrl == df_po.loc[:,"material"])]
            po_next_bo =po_next_bo.sort_values('act_date').reset_index().drop("index",axis=1)
            if( len(po_next_bo) >0):
                df_result1.loc[index,"po_date"]=po_next_bo.loc[0,"act_date"]
                df_result1.loc[index,"poasn_qty"]=po_next_bo.loc[0,"poasn_qty"]


        # if bo bf po no -> days bf po :0
        # yes -> days bf po : if seq ==-1 -> 

        df_result1['#BOdays_bf_pdt']=df_result1.apply(lambda row: 
        max( min(((todays.date() + timedelta(days=row['pdt'])) - row['StartDate'].date()).days+4 , 
                    row['#ofBOdays']),
        0), axis=1)    

        df_result1['#BOdays_bf_pdt'] =df_result1.apply(lambda row: 0 if row.bo_bf_pdt=="no" else row['#BOdays_bf_pdt'], axis=1)

        df_result1=df_result1.rename(columns ={'pdt':'adj. pdt'})
        df_result1['loc']='simulation'
        df_result1.loc[df_result1["BOseq"]!=-1].to_csv(result_loc, index=False)
    end = time.time()
    timelist.append([end-start, "caluculate BO.csv"])

    # %%
    start = time.time()

    def find1(df):
        for i in range(len(df)):
            if df[i]==1:
                return i
        return 0

    df_sumBOseq = df_result.groupby('mtrl').agg({'BOseq': ['sum','count']})
    df_sumBOseq=df_sumBOseq.reset_index()
    df_sumBOseq.columns = ['mtrl', 'BOseq', 'count']
    df_sumBOseq['mtrl_last_index']=(df_sumBOseq['count']).cumsum()-1
    df_sumBOseq['loc1']=df_result.groupby('mtrl')['BOseq'].agg(lambda x: find1(list(x))).reset_index()['BOseq']

    # absolute location of (BOseq==1) for each mtrl = mtrl_last_index-count+(loc1-1)

    df_sumBOseq["StartDate"] = ''   # df_sumBOseq[index][5]
    df_sumBOseq["ox"] = ""          # df_sumBOseq[index][6]

    df_mtrl = pd.DataFrame(df_total["mtrl"].unique())
    df_date = pd.DataFrame(df_total["TheDate"].unique())

    len_mtrl=len(df_mtrl)
    len_date=len(df_date)

    colnames = df_sumBOseq.columns
    df_sumBOseq=df_sumBOseq.to_numpy()

    # for index, row in df_sumBOseq.iterrows():
    for index in range(len(df_sumBOseq)):
        # if row.BOseq > 0:
        if df_sumBOseq[index][1]>0:
            # id = (df_result["mtrl"] == df_sumBOseq[index][0]) & (df_result["BOseq"] == 1)
            # absolute location of (BOseq==1) for each mtrl = mtrl_last_index-count+(loc1-1)        
            id = df_sumBOseq[index][3]-df_sumBOseq[index][4]+1
            # df_sumBOseq[index][5] = df_result.loc[id,"StartDate"].values[0]
            df_sumBOseq[index][5] = df_result.loc[id,"StartDate"]
            df_sumBOseq[index][6] = 'Y'
        # elif row.BOseq< 0:
        elif df_sumBOseq[index][1] < 0:
            df_sumBOseq[index][5] = today 
            df_sumBOseq[index][6] = 'Y'
        # elif row.BOseq == 0:
        elif df_sumBOseq[index][1] == 0:
            # lastday = df_total.loc[df_total.mtrl == df_sumBOseq[index][0]].iloc[-1]
            lastday = df_total.loc[len_date*(index+1)-1]
            # fcst=lastday.fcstD
            # if sum(df_total.loc[df_total.mtrl == df_sumBOseq[index][0], "fcstD"]) == 0:
            if sum(df_total.loc[len_date*index:len_date*(index+1)-1,'fcstD']) == 0:
                # inventory>0 but no fcst
                df_sumBOseq[index][5] = '2100-01-01'
                df_sumBOseq[index][6] = 'N'
            else:
                if lastday.fcstD == 0:
                    fcsts=df_total.loc[len_date*index:len_date*(index+1)-1,'fcstD']
                    # fcst = np.average(df_total.loc[(df_total.mtrl == df_sumBOseq[index][0]) & (df_total.fcstD > 0), "fcstD"].values)
                    if sum(fcsts>0)== 0:
                        fcst=0
                        deltaD=1000
                    fcst = np.average(fcsts[fcsts>0])
                else:  # it means lastday.fcstD>0
                    fcst = lastday.fcstD
                deltaD = lastday.residue/fcst
                if deltaD>1000:
                    deltaD=1000
                elif pd.isnull(deltaD):
                    print(lastday)
                    print(fcst)
                    deltaD=1000
                bo = datetime.strptime(
                    lastday.TheDate, '%Y-%m-%d')+timedelta(days=deltaD)
                df_sumBOseq[index][5] = datetime.strftime(
                    bo, '%Y-%m-%d')
                df_sumBOseq[index][6] = 'N'
        else:
            print(df_sumBOseq[index])
            print("debug needed")

    df_sumBOseq= pd.DataFrame(df_sumBOseq)
    df_sumBOseq.columns= colnames
    df_sumBOseq
    # %%
    BOdateloc = file_loc+"\\"+today+"_"+targetPlant+"_BOdate.csv"
    df_sumBOseq['loc']='simulation'
    df_sumBOseq['days_from_today']=(pd.to_datetime(df_sumBOseq['StartDate']) - datetime.now()).dt.days+1
    df_sumBOseq["DM"]=df_sumBOseq['days_from_today']/365.25*12

    df_sumBOseq[["mtrl", "StartDate",'DM', 'loc']].to_csv(BOdateloc, index=False)
    # %%
    # simulator
    df_simulation= df_sumBOseq.merge(inputKDC[inputKDC["plant"]%100==0], left_on=['mtrl'],right_on=['material'])
    df_simulation= df_simulation[["material","plant","qty","ox","orderlimit","bom",'ms']]
    df_simulation.insert(6,"availability","")
    df_simulation.insert(7,"eta","")

    for index,row in df_simulation.iterrows():
        if(row.bom==1):
            df_simulation.loc[index,"availability"]="check"
            df_simulation.loc[index,"eta"]         ='parent_mtrl'
        elif(row.orderlimit==1):
            df_simulation.loc[index,"availability"]="check"
            df_simulation.loc[index,"eta"]         ='orderlimit'        
        elif(row.ox=="N"):
                df_simulation.loc[index,"availability"]="OK"
        else:
            if sum(df_result1.loc[df_result1["mtrl"]==row.material,"#ofBOdays"])<7:
                df_simulation.loc[index,"availability"]="YES"
                startdate=df_result1.loc[(df_result1["mtrl"]==row.material) & (df_result1["BOseq"]==1),"StartDate"].values[0]
                date=str(startdate).split('T00')[0]+'_'
                df_simulation.loc[index,"eta"]         =date+ str(sum(df_result1.loc[df_result1["mtrl"]==row.material,"#ofBOdays"]))
            else:
                df_simulation.loc[index,"availability"]="NO"
                if len(df_result1.loc[(df_result1["mtrl"]==row.material) & (df_result1["BOseq"]==1),"StartDate"])!=0:
                    startdate=df_result1.loc[(df_result1["mtrl"]==row.material) & (df_result1["BOseq"]==1),"StartDate"].values[0]
                else:
                    startdate=df_result1.loc[(df_result1["mtrl"]==row.material) & (df_result1["BOseq"]==-1),"StartDate"].values[0]
                date=str(startdate).split('T00')[0]+'_'
                df_simulation.loc[index,"eta"]         =date+ str(sum(df_result1.loc[df_result1["mtrl"]==row.material,"#ofBOdays"]))      
        if((row.ms==91) or (row.ms==41)):
            df_simulation.loc[index,"eta"]         =df_simulation.loc[index,"eta"]+'ms'+str(row.ms)
    df_simulation= df_simulation[["material","plant","qty","availability","eta",'ms']]
    # save simulator
    if len(df_result1[df_result.BOseq != 0]) > 0:
        df_simulation=df_simulation.merge(df_result1.loc[df_result1['BOseq']==1,['mtrl','bo_bf_pdt','po_date','poasn_qty','#BOdays_bf_pdt']], how='left',right_on='mtrl',left_on='material').drop('mtrl',axis=1)
    # if targetPlant=='simulate_KDC':
    #     df_simulation_KDC=df_simulation.copy()
    #     df_result1_KDC=df_result1.copy()
    # elif targetPlant=='simulate_LA':
    #     df_simulation_LA=df_simulation.copy()
    #     df_result1_LA=df_result1.copy()
    simul_loc = file_loc+"\\"+today+"_"+targetPlant+"_simulation.csv"
    df_simulation.to_csv(simul_loc,index=False)

    return df_simulation, df_result1

    # end simulator

    # %%
    total_loc = file_loc+"\\"+today+"_"+targetPlant+"_ESA.csv"
    df_total = df_total[['mtrl', 'TheDate', 'On_hand_qty',
                        'residue', 'fcstD',  'BOqty', 'BO$', 'BOseq','loc']].copy()
    df_total.to_csv(total_loc, index=False)
    print('exporting BOdate.csv was done')

    end = time.time()
    timelist.append([end-start, "caluculate BOdate.csv"])

# end simulate_KDC_LA function
if(len(inputKDC)>0):
    targetPlant='simulate_KDC'
    df_simulation_KDC, df_result1_KDC= simulate_KDC_LA(targetPlant)
print('KDC end')
# end NY
################################################################
if(len(inputLA)>0):
    targetPlant='simulate_LA'
    df_simulation_LA, df_result1_LA= simulate_KDC_LA(targetPlant)

# %% LA end
################################################################
print('LA end')
# %% 
replace_material='\''+'\',\''.join(map(str,list(input.material)))+'\''
sql_string="""SELECT material, pdt FROM [ivy.mm.dim.mtrl] WHERE material in ('change_string')
"""

sql_string1=sql_string.replace("'change_string'",replace_material)
df_pdt = pd.read_sql(sql_string1, con=engine)

replace_material='\''+'\',\''.join(map(str,list(input.material)))+'\''
sql_string="""SELECT material, pl_plant as plant, total_stock FROM [ivy.mm.dim.mrp01] WHERE material in ('change_string')
"""

sql_string1=sql_string.replace("'change_string'",replace_material)
df_mrp01 = pd.read_sql(sql_string1, con=engine)

end = time.time()
timelist.append([end-start, "Get pdt"])
# %%

if "order_number" in input.columns:
    # stockcheck
    resultLoc=r"C:\Users\KISS Admin\Desktop\stock check practice"
    simul_loc = resultLoc+"\\"+str(input.order_number.values[0])+"_ivy.xlsx"
    simul_loc1 = resultLoc+"\\"+str(input.order_number.values[0])+"_bo.csv"
else:
    # simulation
    resultLoc=r"C:\Users\KISS Admin\Desktop\stock check practice"
    simul_loc = resultLoc+"\\"+today+"_simulation_total.xlsx"
    simul_loc1 = resultLoc+"\\"+today+"_simulation_bo.csv"

if len(inputLA)==0:
    df_result=df_simulation_KDC
    df_result1=df_result1_KDC
elif len(inputKDC)==0:
    df_result=df_simulation_LA
    df_result1=df_result1_LA
else:
    df_result=pd.concat(df_simulation_KDC,df_simulation_LA)
    df_result1=pd.concat(df_result1_KDC,df_result1_LA)

df_result=df_result.merge(df_pdt,how="left")
df_result["today+pdt"]= df_result["pdt"].apply(lambda x: datetime.strftime(datetime.today() + relativedelta(days=x), ('%Y-%m-%d')))          
# df_result["ms"]= df_result["ms"].apply(lambda x: format(x,'2d'))          

df_result["eta2"]=''
for index, row in df_result.iterrows():
    # print(row)
    if row["availability"]=="NO":
        exp_bodate=datetime.strptime(row.eta.split(' ')[0],'%Y-%m-%d')
        pdtafterdate=datetime.strptime(row["today+pdt"],'%Y-%m-%d')
        if (exp_bodate-pdtafterdate).days<0:
            df_result.loc[index,"eta2"]="bo bf pdt"

df_result=input_order.loc[:,["index","material","plant"]].merge(df_result).drop("index",axis=1)
df_result.to_excel(simul_loc,index=False)
df_result1["BOdays/BOqty"]=df_result1["#ofBOdays"]/df_result1["BOqty"]
df_result1.to_csv(simul_loc1,index=False)

# %%
wb = load_workbook(simul_loc) #Change Location - (type 2)
ws = wb.active
max_row = ws.max_row
max_column = ws.max_column

# conditional formatting : availablity
green_format = PatternFill(fgColor = '00CCFFCC', fill_type='solid')
red_format = PatternFill(fgColor = '00FF8080', fill_type='solid')
blue_format = PatternFill(fgColor = '0000FF80', fill_type='solid')
for k in range(1,max_row+1):
    result_value = str(ws.cell(row=k, column=4).value)
    if result_value == "NO":
        ws.cell(row=k, column=4).fill = red_format
        ws.cell(row=k, column=4).font = Font(color = '00800000')
    elif result_value == "OK":
        ws.cell(row=k, column=4).fill = green_format
        ws.cell(row=k, column=4).font = Font(color = '00008000')
    elif result_value == "YES":
        ws.cell(row=k, column=4).fill = blue_format
        ws.cell(row=k, column=4).font = Font(color = '00008000')
    else:
        ws.cell(row=k, column=4).fill = PatternFill(fgColor = '00FFFFFF', fill_type='solid')
for k in range(1,max_row+1): # 1000, 1110
    result_value = str(ws.cell(row=k, column=2).value)
    if result_value == '1000':
        ws.cell(row=k, column=2).fill = red_format
        ws.cell(row=k, column=2).font = Font(color = '00800000')
    elif result_value == '1110':
        ws.cell(row=k, column=2).fill = green_format
        ws.cell(row=k, column=2).font = Font(color = '00008000')
    else:
        ws.cell(row=k, column=2).fill = PatternFill(fgColor = '00FFFFFF', fill_type='solid')
for k in range(1,max_row+1): # NO and yes
    result_value = str(ws.cell(row=k, column=4).value)
    result_value2 = str(ws.cell(row=k, column=7).value)
    if result_value == 'NO' and result_value2=='yes':
        ws.cell(row=k, column=7).fill = red_format
        ws.cell(row=k, column=4).fill = red_format
        ws.cell(row=k, column=7).font = Font(color = '00800000')
    elif result_value == 'NO' and result_value2=='no':
        ws.cell(row=k, column=7).fill = blue_format
        ws.cell(row=k, column=7).font = Font(color = '00800000')
for k in range(1,max_row+1): # ms 91 or 41
    result_value = str(ws.cell(row=k, column=6).value)
    if (result_value == '41') or (result_value=='91'):
        ws.cell(row=k, column=6).fill = blue_format
        ws.cell(row=k, column=7).font = Font(color = '00800000')
    


for column_cells in ws.columns:
    new_column_length = max(len(str(cell.value)) for cell in column_cells)
    new_column_letter = (get_column_letter(column_cells[0].column))
    if new_column_length > 0:
        ws.column_dimensions[new_column_letter].width = new_column_length*1.3

# Border

# for r in range(1,max_row+1):
#     for c in range(1,max_column+1):
#         ws.cell(row=r, column=c).border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))
if "order_number" in input.columns:
    simul_loc = resultLoc+"\\"+"SC"+str(input.order_number.values[0])+"_ivy.xlsx"
else:
    simul_loc = simul_loc
wb.save(simul_loc) #Change Location - (type 2)
wb.close()

print('Stock check completed!')

# %%
df_time = pd.DataFrame(timelist)
df_time.columns = ["time", "desc"]
df_time["ratio"] = df_time["time"].apply(
    lambda x: f'{(x/sum(df_time["time"])*100):.2f}')
df_time.sort_values("time", ascending=False, inplace=True)
# df_time["time"]=df_time["time"].apply(lambda x : f'{x:.2f}')
df_time = df_time[["desc", "time", "ratio"]]
print(df_time)
# %%
