# %% Load Modules
from multiprocessing.spawn import old_main_modules
from pickle import TRUE
import sys
import win32com.client
import subprocess
from datetime import datetime
import time
import shutil
from calendar import month
from datetime import date, timedelta
import pandas as pd
import numpy as np
import pyodbc
import smartsheet
from sqlalchemy import create_engine

from sqlalchemy.engine import URL
from datetime import datetime
from dateutil.relativedelta import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# %%
import json

with open('data.json', 'r') as f:
    data = json.load(f)

# ID와 비밀번호 가져오기
server = data['server']
database = data['database']
username = data['username']
password = data['password']
connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password
connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url)
print("Connection Established:")

todays = datetime.today()
first_days = todays.replace(day=1)
last_days = datetime(todays.year, todays.month, 1) + relativedelta(months=1) + relativedelta(seconds=-1)
days_left = last_days - todays
today = todays.strftime('%Y-%m-%d')
first_day = first_days.strftime('%Y-%m-%d')
last_day = last_days.strftime('%Y-%m-%d')
business_days = np.busday_count(begindates=first_day, enddates=today) #By today
business_days_thismonth = np.busday_count(begindates=first_day, enddates=last_day)
business_days_left = np.busday_count(begindates=today, enddates=last_day) 

# %%

# Get order information from smartsheet
smartsheet_client = smartsheet.Smartsheet('rjjjwNgTfxwAjE5R5YcSKu5OocAMyLAUJa2av')
smartsheet_client.Sheets.get_sheet_as_csv(
  3482775064995716,           # sheet_id
  r'C:\Users\KISS Admin\Desktop\stock check practice') # TODO RED 

order_df = pd.read_csv(r'C:\Users\KISS Admin\Desktop\stock check practice\Stock Check Request.csv')
order_df['Requester'] = order_df['Requester'].astype(str)

# %%
# downloadLoc = r"C:\Users\KISS Admin\Downloads"
# requestFile = downloadLoc +r"\Stock Check Request.xlsx"

# #order_df = pd.read_csv(r'C:\Users\KISS Admin\Desktop\IVYENT_DH\8. stock check automation code\Stock Check Request.csv')
# order_df = pd.read_excel(requestFile)
# order_df['Requester'] = order_df['Requester'].astype(str)

def salesorg(df):
    if df['Request_team'] =='AST':
        return '1300'
    else:
        return '1100'

order_df['salesorg'] = order_df.apply(salesorg, axis=1)
# the first uncompleted IVY order
# TODO RED
condition = ((order_df['Company (IVY/RED)'] == 'RED') | (order_df['Company (IVY/RED)'] == 'IVY & RED')) & (order_df['Completed'] != True)
order_df = order_df[condition]
order_df = order_df.reset_index()

# %%
if len(order_df) > 0:
    order_df = order_df.iloc[0] # TODO : change index to choose order
    sold_to_party = order_df['Account # (Sold-to Party)']
    po_num = order_df['PO #']
    po_start = order_df['PO Start Date']
    po_end = order_df['PO End Date']
    salesorg = order_df['salesorg']
    order_number=str(order_df["Order #"]).split('.')[0]

    if salesorg == '1300': #For AST orders, we do not check for plant 1000
        plant_list = ['1400', '1410','G140']
    # TODO chsnge RED plane
    
    else:
        plant_list = ['1000', '1400', '1410','G140']
        # TODO chsnge RED plane
    # Data Extraction From SAP

    shell = win32com.client.Dispatch("WScript.Shell")

    # Log in SAP QA for now / will be changed to R3 with system=P01
    # subprocess.check_call(['C:\Program Files (x86)\SAP\FrontEnd\SAPgui\\sapshcut.exe', '-system=P01', '-client=100', '-user=IVY_SADM_05', '-pw=asdf1234', 
    # subprocess.check_call(['C:\Program Files (x86)\SAP\FrontEnd\SAPgui\\sapshcut.exe', '-system=P01', '-client=100', '-user=IVY_SADM_08', '-pw=asdf12348', 
    subprocess.check_call(['C:\Program Files (x86)\SAP\FrontEnd\SAPgui\\sapshcut.exe', '-system=P01', '-client=100', '-user=IVY_SADM_09', '-pw=qwer1234', 
    #  '-command=va05', 
     '-command=ZPPRMRP01', 
    '-type=Transaction', '-max'])
# TODO RED  change sap ID pw
    # Wait until win32com detects SAPGUI
    time.sleep(5)

    # Run vbs code
    def main():
        try:
            SapGuiAuto = win32com.client.GetObject("SAPGUI")
            if not isinstance(SapGuiAuto, win32com.client.CDispatch):
                return

            application = SapGuiAuto.GetScriptingEngine
            if not isinstance(application, win32com.client.CDispatch):
                SapGuiAuto = None
                return

            connection = application.Children(0)
            if not isinstance(connection, win32com.client.CDispatch):
                application = None
                SapGuiAuto = None
                return

            session = connection.Children(0)
            if not isinstance(session, win32com.client.CDispatch):
                connection = None
                application = None
                SapGuiAuto = None
                return

            # SapGuiAuto = win32com.client.GetObject("SAPGUI")
            # application = SapGuiAuto.GetScriptingEngine
            # connection = application.Children(0)
            # session = connection.Children(0)

            session.findById("wnd[0]").maximize
            session.findById("wnd[0]/tbar[0]/okcd").text = "/nva05"
            session.findById("wnd[0]").sendVKey (0)
            session.findById("wnd[0]/usr/ctxtVBCOM-KUNDE").text = sold_to_party #soldtoparty
            session.findById("wnd[0]/usr/txtVBCOM-BSTKD").text = po_num #PO number
            session.findById("wnd[0]/usr/ctxtVBCOM-AUDAT").text = po_start #PO document date (start)
            session.findById("wnd[0]/usr/ctxtVBCOM-AUDAT_BIS").text = po_end #PO document date (end)
            session.findById("wnd[0]").sendVKey (0)
            session.findById("wnd[1]/usr/ctxtVBCOM-VKORG").text = salesorg #salesorg : ivy 1100 / red 1400 / AST 1300
            session.findById("wnd[1]/usr/ctxtVBCOM-VKORG").caretPosition = 4
            session.findById("wnd[1]").sendVKey (0)
            session.findById("wnd[0]/tbar[1]/btn[32]").press()
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER2_LAYO/shellcont/shell").selectedRows = "0-18"
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER2_LAYO/shellcont/shell").doubleClickCurrentCell()
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = (13)
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = (9)
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "13"
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = (11)
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "11"
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = (86)
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = (80)
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "86"
            session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").contextMenu()
            session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").selectContextMenuItem ("&XXL")
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
#            session.findById("wnd[1]/usr/ctxtDY_PATH").text = "Y:\\OM ONLY_Shared Documents\\3 Teams\\2 DP\\38. Stockcheck" #Change Location - (type 1)
            # session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"C:\Users\KISS Admin\Downloads" #Change Location - (type 1)
            # session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "input.XLSX"
            session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 7
            session.findById("wnd[1]/tbar[0]/btn[11]").press()
            print("SAP good")
        
        except:
            print(sys.exc_info()[0])

        finally:
            session = None
            connection = None
            application = None
            SapGuiAuto = None


    if __name__ == "__main__":
        main()

    '''pykill.py - selective process kill prog'''
    import psutil

    def main():
  
        for proc in psutil.process_iter():
        # check whether the process name matches
        # print(proc.name())
            if any(procstr in proc.name() for procstr in\
                ['saplogon', 'EXCEL']):
                print(f'Killing {proc.name()}')
                proc.kill()


    if __name__ == "__main__":
        main()

    # %% Import Input file
    import os
    # print(order_number)

    old_name=r"C:\Users\KISS Admin\Documents\SAP\SAP GUI\export.XLSX"
    # old_name=r"C:\Users\KISS Admin\Downloads\input.XLSX"
    # new_name=r"C:\Users\KISS Admin\Desktop\IVYENT_DH\7. stock check"+"\\"+order_number+"_input.XLSX"
    # new_name=r"C:\Users\KISS Admin\Desktop\stock check practice\input.XLSX"
    # os.rename(old_name,new_name)


    input_df = pd.read_excel(old_name, sheet_name='Sheet1') #Change Location - (type 1)
    input_df = input_df[['Material', 'Order Quantity', 'Plant']]
    input_df.columns = ['material', 'qty', 'plant']
    input_df['plant'] = input_df['plant'].astype(str)
    input_df['material'] = input_df['material'].astype(str)

    # %% Listup : Order limit & BO products
    # # orderlimit_df : Order limit list
    # orderlimit_df = pd.read_sql("""Select material, from_date, to_date From [ivy.mm.dim.orderlimit]""", con=engine)
    orderlimit_df = pd.read_sql("""SELECT material, from_date, to_date FROM [ivy.mm.dim.orderlimit] WHERE from_date<=GETDATE() and to_date>=GETDATE()""", con=engine)
    orderlimit_df.columns = ['material', 'from_date', 'to_date']
    orderlimit_df['orderlimit'] = 1
    orderlimit_df = orderlimit_df.drop_duplicates(subset='material')

    # # bo_df : BackOrder list
    # bo_df = pd.read_sql("""Select material from [ivy.sd.fact.listofbo]""", con=engine)
    # bo_df.columns = ['material']
    # bo_df['bo'] = 1


    # Current stock (dimmrp01)

    stock_df = pd.read_sql("""select material, pl_plant as plant, ms, (total_stock - blocked) as stock, safetystock, avg_rqmts as requirements, pdt, openorder, opendeliv
    from [ivy.mm.dim.mrp01]
    """, con=engine)
    stock_df['plant'] = stock_df['plant'].astype(str)
    stock_df['requirements'] = stock_df['requirements'] * 53 # TODO Why
    # %%
    # FCST
    fcst1_df = pd.read_sql("""select material, plant, sum(eship) as fcst1 from [ivy.mm.dim.factfcst]
    where year(act_date) = year(getdate())
    and month(act_date) = month(getdate())
    and cg_key = 'TR'
    group by material, plant""", con=engine)
    fcst1_df['fcst_thismonth'] = fcst1_df['fcst1'] / (last_days - first_days).days * days_left.days # Calendarday
    # %%
    fcst2_df = pd.read_sql("""select material, plant, sum(eship) as fcst2 from [ivy.mm.dim.factfcst]
    where year(act_date) = year(getdate()) 
    and month(act_date) between month(getdate())+1 and month(getdate()) + 2
    and cg_key = 'TR'
    group by material, plant""", con=engine)
    fcst2_df['fcst_nextmonths'] = fcst2_df['fcst2'] / 60 * (53-int(days_left.days))

    fcst_df = pd.merge(fcst1_df, fcst2_df, how='inner', on=['material', 'plant'])
    fcst_df['fcst'] = fcst_df['fcst_thismonth'] + fcst_df['fcst_nextmonths']
    fcst_df = fcst_df[['material', 'plant', 'fcst']]
    fcst_df['plant'] = fcst_df['plant'].astype(str)
    fcst_df['fcst'] = fcst_df['fcst'].astype(int)

    # %% PO table
    po_df = pd.read_sql("""Select material, plant, min(act_date) as eta, (sum(po_qty) + sum(asn_qty)) as po_qty
    From [ivy.mm.dim.fact_poasn]
    Where act_date between getdate() and getdate() + 53 
    and plant in ('1000', '1100', '1110', '1400', '1410','G140')
    Group by material, plant""", con=engine) # assume 7 days of PO delay
    po_df['plant'] = po_df['plant'].astype(str)

    # %% BOM (dimbom)

    bom_df = pd.read_sql("""select bom_parent_material as material from [ivy.mm.dim.bom_aset]""", con=engine)
    bom_df['bom'] = 1 

    # %% Final_df : master table
    merge1_df = pd.merge(input_df, stock_df, on = ['material', 'plant'], how = 'left')
    merge2_df = pd.merge(merge1_df, po_df, on = ['material', 'plant'], how = 'left')
    merge3_df = pd.merge(merge2_df, fcst_df, on = ['material', 'plant'], how = 'left')
    merge4_df = pd.merge(merge3_df, orderlimit_df, on='material', how='left') #if order limit, then orderlimit column == 1
    #merge5_df = pd.merge(merge4_df, bo_df, on='material', how='left') #if bo, then bo column == 1
    merge6_df = pd.merge(merge4_df, bom_df, on='material', how='left') #if bom, then bom column == 1

    # %%
    final_df = merge6_df[['material', 'plant', 'ms', 'qty', 'requirements', 'fcst', 'eta', 'stock', 'orderlimit', 'bom', 'po_qty', 'safetystock', 'pdt', 'openorder', 'opendeliv']]
    final_df.columns = ['material', 'plant', 'ms', 'qty', 'requirements', 'fcst', 'eta', 'stock', 'orderlimit', 'bom', 'po_qty', 'safetystock', 'pdt', 'openorder', 'opendeliv']
    final_df['availability']='None'
    final_df[['orderlimit','bom', 'stock', 'po_qty', 'safetystock', 'pdt', 'fcst', 'openorder', 'opendeliv']] = final_df[['orderlimit','bom', 'stock', 'po_qty', 'safetystock', 'pdt', 'fcst', 'openorder', 'opendeliv']].fillna(0) # select specific column
    final_df['orderlimit'] = final_df['orderlimit'].astype('int')
    #final_df['bo'] = final_df['bo'].astype('int')
    final_df['bom'] = final_df['bom'].astype('int')
    final_df.reset_index(inplace=True)
    final_df.drop(['index'], axis=1, inplace=True)
    final_df['eta'] = final_df['eta'].fillna('NO PLAN')

    # %%
    #Ivy
    final_df = final_df[final_df['plant'].isin(plant_list)]

    # Logic : BOM? Order Limit?
    def limit_check(final_df):
        if (final_df['ms'] == '41') or (final_df['ms'] == '91'):
            return 'ms 41 or 91'
        elif (final_df['bom'] == 1) or ("ASET" in final_df['material']):
            return 'Review'
        elif (final_df['orderlimit'] == 1):
            return 'NO'
        else:
            return 'OK'

    # %% 
    # Logic : Expected BO?

    def bo_check(final_df):
        if (final_df['orderlimit'] != 1) and (final_df['availability'] != 'Review'):
            if final_df['plant'] == '1000': #For plant 1000, please double check. Some items do not have PO info
                if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                    return 'OK'
                else:
                    return 'NO'
                        
            elif final_df['plant'] == '1100':
                if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                    return 'OK'
                else:
                    return 'NO'

            elif final_df['plant'] == '1110':
                if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                    return 'OK'
                else:
                    return 'NO'
        
            elif final_df['plant'] == '1400':
                if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                    return 'OK'
                else:
                    return 'NO'
            
            elif final_df['plant'] == 'G140':
                if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                    return 'OK'
                else:
                    return 'NO'
        
    
            else:
                if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                    return 'OK'
                else:
                    return 'NO'
        
        elif final_df['orderlimit'] == 1:
            return 'NO'
    
        # elif final_df['bo'] == 1:
        #     return 'NO'
        
        else:
            return 'Review'

    final_df['availability'] = final_df.apply(limit_check, axis=1)
    final_df['availability'] = final_df.apply(bo_check, axis=1)
             
    # Clear eta if availability == 'OK'
    def eta(final_df):
        if (final_df['availability'] == 'OK') or (final_df['availability'] == 'Review'):
            return ""
        elif final_df['orderlimit'] == 1:
            return "Order Limit"
        elif final_df['eta'] == 0:
            return ""
        else:
            return final_df['eta']

    final_df['eta'] = final_df.apply(eta, axis=1)
    print(final_df)
    
    # %% Finalize

    final_df = final_df[['material', 'qty', 'plant', 'availability', 'eta','ms']]
    print(final_df)

    # %%
    # Export output
    resultLoc=r"C:\Users\KISS Admin\Desktop\stock check practice"
    final_df.to_excel(resultLoc+"\\"+order_number+"Result_RED.xlsx") #Change Location - (type 2)

    # %% Excel file formatting
    wb = load_workbook(resultLoc+"\\"+order_number+"Result_IVY.xlsx") #Change Location - (type 2)
    ws = wb.active
    max_row = ws.max_row
    max_column = ws.max_column

    # conditional formatting : availablity
    green_format = PatternFill(fgColor = '00CCFFCC', fill_type='solid')
    red_format = PatternFill(fgColor = '00FF8080', fill_type='solid')
    for k in range(1,max_row+1):
        result_value = str(ws.cell(row=k, column=5).value)
        if result_value == "NO":
            ws.cell(row=k, column=5).fill = red_format
            ws.cell(row=k, column=5).font = Font(color = '00800000')
        elif result_value == "OK":
            ws.cell(row=k, column=5).fill = green_format
            ws.cell(row=k, column=5).font = Font(color = '00008000')
        else:
            ws.cell(row=k, column=5).fill = PatternFill(fgColor = '00FFFFFF', fill_type='solid')

    # Border

    for r in range(1,max_row+1):
        for c in range(1,max_column+1):
            ws.cell(row=r, column=c).border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))

    wb.save(resultLoc+"\\"+order_number+"Result_RED.xlsx") #Change Location - (type 2)
    wb.close()

    print('Stock check completed!')
    # new_name=r"C:\Users\KISS Admin\Desktop\stock check practice\input.XLSX"
    # os.rename(old_name,new_name)
    os.remove(old_name)
    os.remove(r'C:\Users\KISS Admin\Desktop\stock check practice\Stock Check Request.csv')
else:
    print('There is no stock check request!')


# %%
