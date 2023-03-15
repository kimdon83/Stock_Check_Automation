# %% Load Modules
from multiprocessing.spawn import old_main_modules
from pickle import TRUE
import sys
import win32com.client
import subprocess
from datetime import datetime
import time
import shutil
from calendar import month
from datetime import date, timedelta
import pandas as pd
import numpy as np
import pyodbc
import smartsheet
from sqlalchemy import create_engine

from sqlalchemy.engine import URL
from datetime import datetime
from dateutil.relativedelta import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# %%
server = '10.1.3.25' 
database = 'KIRA' 
username = 'kiradba'                                                                                                                                                                                                                  
password = 'Kiss!234!' 
connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password
connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url)
print("Connection Established:")

todays = datetime.today()
first_days = todays.replace(day=1)
last_days = datetime(todays.year, todays.month, 1) + relativedelta(months=1) + relativedelta(seconds=-1)
days_left = last_days - todays
today = todays.strftime('%Y-%m-%d')
first_day = first_days.strftime('%Y-%m-%d')
last_day = last_days.strftime('%Y-%m-%d')
business_days = np.busday_count(begindates=first_day, enddates=today) #By today
business_days_thismonth = np.busday_count(begindates=first_day, enddates=last_day)
business_days_left = np.busday_count(begindates=today, enddates=last_day) 

# %%
# %% Import Input file
import os
# print(order_number)
old_name=r"C:\Users\KISS Admin\Downloads\input.XLSX"
new_name=r"C:\Users\KISS Admin\Desktop\IVYENT_DH\7. stock check"+"\\"+order_number+"_input.XLSX"
os.rename(old_name,new_name)

#     if salesorg == '1300': #For AST orders, we do not check for plant 1000
#         plant_list = ['1100', '1110']
    
#     else:
#         plant_list = ['1000', '1100', '1110']


input_df = pd.read_excel(new_name, sheet_name='Sheet1') #Change Location - (type 1)
input_df = input_df[['Material', 'Order Quantity', 'Plant']]
input_df.columns = ['material', 'qty', 'plant']
input_df['plant'] = input_df['plant'].astype(str)
input_df['material'] = input_df['material'].astype(str)

# %% Listup : Order limit & BO products
# # orderlimit_df : Order limit list
orderlimit_df = pd.read_sql("""Select material, from_date, to_date From [ivy.mm.dim.orderlimit]""", con=engine)
orderlimit_df.columns = ['material', 'from_date', 'to_date']
orderlimit_df['orderlimit'] = 1
orderlimit_df = orderlimit_df.drop_duplicates(subset='material')

# # bo_df : BackOrder list
# bo_df = pd.read_sql("""Select material from [ivy.sd.fact.listofbo]""", con=engine)
# bo_df.columns = ['material']
# bo_df['bo'] = 1

# Current stock (dimmrp01)

stock_df = pd.read_sql("""select material, pl_plant as plant, ms, (total_stock - blocked) as stock, safetystock, avg_rqmts as requirements, pdt, openorder, opendeliv
from [ivy.mm.dim.mrp01]
""", con=engine)
stock_df['plant'] = stock_df['plant'].astype(str)
stock_df['requirements'] = stock_df['requirements'] * 53 # TODO Why
# %%
# FCST
fcst1_df = pd.read_sql("""select material, plant, sum(eship) as fcst1 from [ivy.mm.dim.factfcst]
where year(act_date) = year(getdate())
and month(act_date) = month(getdate())
and cg_key = 'TR'
group by material, plant""", con=engine)
fcst1_df['fcst_thismonth'] = fcst1_df['fcst1'] / (last_days - first_days).days * days_left.days # Calendarday
# %%
fcst2_df = pd.read_sql("""select material, plant, sum(eship) as fcst2 from [ivy.mm.dim.factfcst]
where year(act_date) = year(getdate()) 
and month(act_date) between month(getdate())+1 and month(getdate()) + 2
and cg_key = 'TR'
group by material, plant""", con=engine)
fcst2_df['fcst_nextmonths'] = fcst2_df['fcst2'] / 60 * (53-int(days_left.days))

fcst_df = pd.merge(fcst1_df, fcst2_df, how='inner', on=['material', 'plant'])
fcst_df['fcst'] = fcst_df['fcst_thismonth'] + fcst_df['fcst_nextmonths']
fcst_df = fcst_df[['material', 'plant', 'fcst']]
fcst_df['plant'] = fcst_df['plant'].astype(str)
fcst_df['fcst'] = fcst_df['fcst'].astype(int)

# %% PO table
po_df = pd.read_sql("""Select material, plant, min(act_date) as eta, (sum(po_qty) + sum(asn_qty)) as po_qty
From [ivy.mm.dim.fact_poasn]
Where act_date between getdate() and getdate() + 53 
and plant in ('1000', '1100', '1110', '1400', '1410')
Group by material, plant""", con=engine) # assume 7 days of PO delay
po_df['plant'] = po_df['plant'].astype(str)

# %% BOM (dimbom)

bom_df = pd.read_sql("""select bom_parent_material as material from [ivy.mm.dim.bom]""", con=engine)
bom_df['bom'] = 1 

# %% Final_df : master table
merge1_df = pd.merge(input_df, stock_df, on = ['material', 'plant'], how = 'left')
merge2_df = pd.merge(merge1_df, po_df, on = ['material', 'plant'], how = 'left')
merge3_df = pd.merge(merge2_df, fcst_df, on = ['material', 'plant'], how = 'left')
merge4_df = pd.merge(merge3_df, orderlimit_df, on='material', how='left') #if order limit, then orderlimit column == 1
#merge5_df = pd.merge(merge4_df, bo_df, on='material', how='left') #if bo, then bo column == 1
merge6_df = pd.merge(merge4_df, bom_df, on='material', how='left') #if bom, then bom column == 1

# %%
final_df = merge6_df[['material', 'plant', 'ms', 'qty', 'requirements', 'fcst', 'eta', 'stock', 'orderlimit', 'bom', 'po_qty', 'safetystock', 'pdt', 'openorder', 'opendeliv']]
final_df.columns = ['material', 'plant', 'ms', 'qty', 'requirements', 'fcst', 'eta', 'stock', 'orderlimit', 'bom', 'po_qty', 'safetystock', 'pdt', 'openorder', 'opendeliv']
final_df['availability']='None'
final_df[['orderlimit','bom', 'stock', 'po_qty', 'safetystock', 'pdt', 'fcst', 'openorder', 'opendeliv']] = final_df[['orderlimit','bom', 'stock', 'po_qty', 'safetystock', 'pdt', 'fcst', 'openorder', 'opendeliv']].fillna(0) # select specific column
final_df['orderlimit'] = final_df['orderlimit'].astype('int')
#final_df['bo'] = final_df['bo'].astype('int')
final_df['bom'] = final_df['bom'].astype('int')
final_df.reset_index(inplace=True)
final_df.drop(['index'], axis=1, inplace=True)
final_df['eta'] = final_df['eta'].fillna('NO PLAN')

# %%
#Ivy
final_df = final_df[final_df['plant'].isin(plant_list)]

# Logic : BOM? Order Limit?
def limit_check(final_df):
    if (final_df['ms'] == '41') or (final_df['ms'] == '91'):
        return 'ms 41 or 91'
    elif (final_df['bom'] == 1) or ("ASET" in final_df['material']):
        return 'Review'
    elif (final_df['orderlimit'] == 1):
        return 'NO'
    else:
        return 'OK'

# %% 
# Logic : Expected BO?

def bo_check(final_df):
    if (final_df['orderlimit'] != 1) and (final_df['availability'] != 'Review'):
        if final_df['plant'] == '1000': #For plant 1000, please double check. Some items do not have PO info
            if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                return 'OK'
            else:
                return 'NO'
                    
        elif final_df['plant'] == '1100':
            if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                return 'OK'
            else:
                return 'NO'

        elif final_df['plant'] == '1110':
            if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                return 'OK'
            else:
                return 'NO'
    
        elif final_df['plant'] == '1400':
            if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                return 'OK'
            else:
                return 'NO'
    

        else:
            if final_df['stock'] - final_df['qty'] - max(final_df['requirements'],final_df['fcst']+final_df['openorder']+final_df['opendeliv']) + final_df['po_qty'] >= 0:
                return 'OK'
            else:
                return 'NO'
    
    elif final_df['orderlimit'] == 1:
        return 'NO'

    # elif final_df['bo'] == 1:
    #     return 'NO'
    
    else:
        return 'Review'

final_df['availability'] = final_df.apply(limit_check, axis=1)
final_df['availability'] = final_df.apply(bo_check, axis=1)
            
# Clear eta if availability == 'OK'
def eta(final_df):
    if (final_df['availability'] == 'OK') or (final_df['availability'] == 'Review'):
        return ""
    elif final_df['orderlimit'] == 1:
        return "Order Limit"
    elif final_df['eta'] == 0:
        return ""
    else:
        return final_df['eta']

final_df['eta'] = final_df.apply(eta, axis=1)
print(final_df)

# %% Finalize

final_df = final_df[['material', 'qty', 'plant', 'availability', 'eta']]
print(final_df)

# %%
# Export output
resultLoc=r"C:\Users\KISS Admin\Desktop\IVYENT_DH\7. stock check"
final_df.to_excel(resultLoc+"\\"+order_number+"SCresult_IVY.xlsx") #Change Location - (type 2)

# %% Excel file formatting
wb = load_workbook(resultLoc+"\\"+order_number+"SCresult_IVY.xlsx") #Change Location - (type 2)
ws = wb.active
max_row = ws.max_row
max_column = ws.max_column

# conditional formatting : availablity
green_format = PatternFill(fgColor = '00CCFFCC', fill_type='solid')
red_format = PatternFill(fgColor = '00FF8080', fill_type='solid')
for k in range(1,max_row+1):
    result_value = str(ws.cell(row=k, column=5).value)
    if result_value == "NO":
        ws.cell(row=k, column=5).fill = red_format
        ws.cell(row=k, column=5).font = Font(color = '00800000')
    elif result_value == "OK":
        ws.cell(row=k, column=5).fill = green_format
        ws.cell(row=k, column=5).font = Font(color = '00008000')
    else:
        ws.cell(row=k, column=5).fill = PatternFill(fgColor = '00FFFFFF', fill_type='solid')

# Border

for r in range(1,max_row+1):
    for c in range(1,max_column+1):
        ws.cell(row=r, column=c).border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))

wb.save(resultLoc+"\\"+order_number+"SCresult_IVY.xlsx") #Change Location - (type 2)
wb.close()

print('Stock check completed!')

# else:
#     print('There is no stock check request!')

#     old_name=r"C:\Users\KISS Admin\Downloads\Stock Check Request.XLSX"
#     new_name=r"C:\Users\KISS Admin\Downloads"+"\\"+order_number+"Stock Check Request.XLSX"
#     os.rename(old_name,new_name)



# %%
