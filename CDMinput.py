import tkinter as tk
from tkinter import filedialog
import os
import pandas as pd
from datetime import datetime
import subprocess
import time
from sqlalchemy import create_engine
from sqlalchemy.engine import URL

def open_excel_file():
    # Default file location
    default_location = r"C:\Users\dokim2\Documents\SOM_20231127_1100.xlsx"

    # Create a GUI window
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Open file dialog
    file_path = filedialog.askopenfilename(initialdir=os.path.dirname(default_location),
                                           title="Select file",
                                           filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))

    root.destroy()  # Close the GUI window
    return file_path

def read_excel_to_dataframe(file_path):
    # Read the file into a pandas DataFrame
    if file_path:
        df = pd.read_excel(file_path)
        return df
    else:
        return None

def rename_columns(dataframe):
    # Rename columns to "material" and "qty"
    if dataframe is not None and len(dataframe.columns) >= 2:
        dataframe.columns = ["material", "qty"] + list(dataframe.columns[2:])
        return dataframe
    else:
        return dataframe

def get_salesorg(file_path):
    # Extract salesorg from file name
    if file_path and os.path.basename(file_path).endswith("1100.xlsx"):
        return "1100"
    else:
        return "1400"
def addPlant_salesorg(dataframe, salesorg,date_str_input):
    if salesorg =='1100':
        dataframe['plant']=salesorg
        dataframe['order_number']=date_str_input
        dataframe['salesorg']=salesorg
    else:
        dataframe['plant']='G140'
        dataframe['order_number']=date_str_input
        dataframe['salesorg']=salesorg


def save_to_input_file(dataframe,loc_input):
    # Delete the existing file if it exists
    if os.path.exists(loc_input):
        os.remove(loc_input)

    dataframe.to_csv(loc_input,index=False)

file_path = open_excel_file()

if file_path:
    dataframe = read_excel_to_dataframe(file_path)
    dataframe = rename_columns(dataframe)
    date_str=datetime.today().date().strftime("%m%d%Y")
    date_str=date_str[1:] if date_str[0]=='0' else date_str
    salesorg = get_salesorg(file_path)
    addPlant_salesorg(dataframe, salesorg,date_str)

    # %%
    # Connect to KIRA server
    start = time.time()

    import json
    desktoploc=r'C:\Users\dokim2\OneDrive - Kiss Products Inc\Desktop'

    with open(desktoploc+'\IVYENT_DH\data.json', 'r') as f:
        data = json.load(f)
    desktoploc=r'C:\Users\dokim2\Documents\Stock Check Result'

    # getID and password
    server = data['server']
    database = data['database']
    username = data['username']
    password = data['password']
    connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + \
        server+';DATABASE='+database+';UID='+username+';PWD=' + password
    connection_url = URL.create(
        "mssql+pyodbc", query={"odbc_connect": connection_string})
    engine = create_engine(connection_url, fast_executemany=True)
    print("Connection Established:")

    end = time.time()

    # %%
    ss_df= pd.read_sql("""SELECT material, pl_plant as plant, safetystock FROM [ivy.mm.dim.mrp01]
WHERE pl_plant in ('1100','G140')""", con=engine)
    
    df_input = dataframe.merge(ss_df, how='left',on=["material",'plant'])
    df_input["safetystock"] = df_input["safetystock"].fillna(0)

    loc_input = r'C:\Users\dokim2\Documents\Stock Check Result'+'\simulator_input.csv'
    # loc_input2 = r'C:\Users\dokim2\Documents\Stock Check Result'+'\simulator_input_keep for CDM.csv'

    ss_ratio = input("input ratio of safety stock as percent. (Press 'Enter' for default 25%) :")

    if ss_ratio=="":
        ss_ratio=0.25
    else:
        ss_ratio = float(ss_ratio)/100

    # save_to_input_file(df_input,loc_input2)
    df_input_backup = df_input.copy()
    df_input["qty"] = df_input["qty"] + df_input["safetystock"].apply(lambda x: int(x * ss_ratio))

    df_input0=df_input.drop("safetystock",axis=1)
    save_to_input_file(df_input0,loc_input)

    print(f"DataFrame:\n{dataframe}\nSalesorg: {salesorg}")

    ## run python code for simulation
    # loc_simulator=r'C:\Users\dokim2\OneDrive - Kiss Products Inc\Desktop\IVYENT_DH\P4. stock check automation code\dailyDM_simulator_beta.py'
    loc_simulator='dailyDM_simulator_beta.py'

    subprocess.run(["python", loc_simulator], check=True)
    desktoploc=r'C:\Users\dokim2\Documents\Stock Check Result'
    if salesorg =='1100':
        simul_loc = desktoploc+"\\"+"SC"+date_str+"_ivy.xlsx"
    else:
        simul_loc = desktoploc+"\\"+"SC"+date_str+"_red.xlsx"

    CDMscResult=read_excel_to_dataframe(simul_loc)
    print(CDMscResult)

    CDMscResult = CDMscResult.rename(columns={'qty': 'qty+SS'})
    merged_df = CDMscResult.merge(df_input_backup[['material', 'qty', 'safetystock']], on='material', how='left')

    # Save the updated DataFrame back to the Excel file
    merged_df.to_excel(simul_loc, index=False)

    # %%

    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(simul_loc) #Change Location - (type 2)
    ws = wb.active
    max_row = ws.max_row
    max_column = ws.max_column

    # conditional formatting : availablity
    green_format = PatternFill(fgColor = '00CCFFCC', fill_type='solid')
    red_format = PatternFill(fgColor = '00FF8080', fill_type='solid')
    blue_format = PatternFill(fgColor = '0000FF80', fill_type='solid')
    for k in range(1,max_row+1):
        result_value = str(ws.cell(row=k, column=4).value)
        if result_value == "NO":
            ws.cell(row=k, column=4).fill = red_format
            ws.cell(row=k, column=4).font = Font(color = '00800000')
        elif result_value == "OK":
            ws.cell(row=k, column=4).fill = green_format
            ws.cell(row=k, column=4).font = Font(color = '00008000')
        elif result_value == "YES":
            ws.cell(row=k, column=4).fill = blue_format
            ws.cell(row=k, column=4).font = Font(color = '00008000')
        else:
            ws.cell(row=k, column=4).fill = PatternFill(fgColor = '00FFFFFF', fill_type='solid')
    for k in range(1,max_row+1): # 1000, 1110
        result_value = str(ws.cell(row=k, column=2).value)
        if result_value == '1000':
            ws.cell(row=k, column=2).fill = red_format
            ws.cell(row=k, column=2).font = Font(color = '00800000')
        elif result_value == '1110':
            ws.cell(row=k, column=2).fill = green_format
            ws.cell(row=k, column=2).font = Font(color = '00008000')
        else:
            ws.cell(row=k, column=2).fill = PatternFill(fgColor = '00FFFFFF', fill_type='solid')
    for k in range(1,max_row+1): # NO and yes
        result_value = str(ws.cell(row=k, column=4).value)
        result_value2 = str(ws.cell(row=k, column=7).value)
        if result_value == 'NO' and result_value2=='yes':
            ws.cell(row=k, column=7).fill = red_format
            ws.cell(row=k, column=4).fill = red_format
            ws.cell(row=k, column=7).font = Font(color = '00800000')
        elif result_value == 'NO' and result_value2=='no':
            ws.cell(row=k, column=7).fill = blue_format
            ws.cell(row=k, column=7).font = Font(color = '00800000')
    for k in range(1,max_row+1): # ms 91 or 41
        result_value = str(ws.cell(row=k, column=6).value)
        if (result_value == '41') or (result_value=='91'):
            ws.cell(row=k, column=6).fill = blue_format
            ws.cell(row=k, column=7).font = Font(color = '00800000')   

    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.3
    
    ws['T1'] = 'adj. qty'
    # Applying the formula to each cell in the T column
    for row in range(2, max_row + 1):  # Starting from row 2 to avoid applying the formula to the header
        ws[f'T{row}'] = f'=MAX(R{row}-K{row}-MOD(R{row}-K{row},O{row}),0)'

    ws.column_dimensions[get_column_letter(5)].width = 20

    ws.auto_filter.ref = f"A1:T{max_row}"

    wb.save(simul_loc) #Change Location - (type 2)
    wb.close()


    print("good?")
