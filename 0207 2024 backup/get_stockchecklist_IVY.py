# %% Load Modules
from multiprocessing.spawn import old_main_modules
from pickle import TRUE
import sys
import win32com.client
import subprocess
from datetime import datetime
import time
import shutil
from calendar import month
from datetime import date, timedelta
import pandas as pd
import numpy as np
import pyodbc
import smartsheet
from sqlalchemy import create_engine

from sqlalchemy.engine import URL
from datetime import datetime
from dateutil.relativedelta import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
# %%
import pyautogui
import pyperclip

import psutil

# %%
import json
driveloc=r'C:\Users\dokim2\OneDrive - Kiss Products Inc'
desktoploc=r'C:\Users\dokim2\OneDrive - Kiss Products Inc\Desktop'
with open(desktoploc+'\IVYENT_DH\data.json', 'r') as f:
    data = json.load(f)

# ID와 비밀번호 가져오기
server = data['server']
database = data['database']
username = data['username']
password = data['password']
connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password
connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url)
print("Connection Established:")

todays = datetime.today()
first_days = todays.replace(day=1)
last_days = datetime(todays.year, todays.month, 1) + relativedelta(months=1) + relativedelta(seconds=-1)
days_left = last_days - todays
today = todays.strftime('%Y-%m-%d')
first_day = first_days.strftime('%Y-%m-%d')
last_day = last_days.strftime('%Y-%m-%d')
business_days = np.busday_count(begindates=first_day, enddates=today) #By today
business_days_thismonth = np.busday_count(begindates=first_day, enddates=last_day)
business_days_left = np.busday_count(begindates=today, enddates=last_day) 

# %%
def pskill():
    for proc in psutil.process_iter():
    # check whether the process name matches
    # print(proc.name())
        if any(procstr in proc.name() for procstr in\
            ['saplogon', 'EXCEL']):
            print(f'Killing {proc.name()}')
            proc.kill()

def salesorg(df):
    if df['Request_team'] =='AST':
        return '1300'
    else:
        return '1100'


def getfromVA03(ordernumber_in):
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        if not isinstance(SapGuiAuto, win32com.client.CDispatch):
            return

        application = SapGuiAuto.GetScriptingEngine
        if not isinstance(application, win32com.client.CDispatch):
            SapGuiAuto = None
            return

        connection = application.Children(0)
        if not isinstance(connection, win32com.client.CDispatch):
            application = None
            SapGuiAuto = None
            return

        session = connection.Children(0)
        if not isinstance(session, win32com.client.CDispatch):
            connection = None
            application = None
            SapGuiAuto = None
            return

        # SapGuiAuto = win32com.client.GetObject("SAPGUI")
        # application = SapGuiAuto.GetScriptingEngine
        # connection = application.Children(0)
        # session = connection.Children(0)

        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nva03"
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/ctxtVBAK-VBELN").text = ordernumber_in
        session.findById("wnd[0]/usr/ctxtVBAK-VBELN").caretPosition = 7
        session.findById("wnd[0]").sendVKey (0)
        time.sleep(1)

        pyautogui.hotkey('enter')

        session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/subPART-SUB:SAPMV45A:4701/ctxtKUAGV-KUNNR").caretPosition = 8

        session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/subPART-SUB:SAPMV45A:4701/ctxtKUAGV-KUNNR").setFocus()
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.2)
        clipboard_value1 = pyperclip.paste()
        # session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/subPART-SUB:SAPMV45A:4701/ctxtKUAGV-KUNNR").caretPosition = 8
        # session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/subPART-SUB:SAPMV45A:4701/ctxtKUAGV-KUNNR").showContextMenu()
        session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/txtVBKD-BSTKD").setFocus()
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.2)
        clipboard_value2 = pyperclip.paste()



        session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/ctxtVBKD-BSTDK").setFocus()
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.2)
        clipboard_value3 = pyperclip.paste()
        # session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/ctxtVBKD-BSTDK").caretPosition = 0

        print("SAP good")
    
    except:
        print(sys.exc_info()[0])

    finally:
        session = None
        connection = None
        application = None
        SapGuiAuto = None

        return clipboard_value1,clipboard_value2,clipboard_value3

def main(sold_to_party, po_num, po_date,salesorg):
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        if not isinstance(SapGuiAuto, win32com.client.CDispatch):
            return

        application = SapGuiAuto.GetScriptingEngine
        if not isinstance(application, win32com.client.CDispatch):
            SapGuiAuto = None
            return

        connection = application.Children(0)
        if not isinstance(connection, win32com.client.CDispatch):
            application = None
            SapGuiAuto = None
            return

        session = connection.Children(0)
        if not isinstance(session, win32com.client.CDispatch):
            connection = None
            application = None
            SapGuiAuto = None
            return

        # SapGuiAuto = win32com.client.GetObject("SAPGUI")
        # application = SapGuiAuto.GetScriptingEngine
        # connection = application.Children(0)
        # session = connection.Children(0)

        session.findById("wnd[0]").maximize
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nva05"
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/ctxtVBCOM-KUNDE").text = sold_to_party #soldtoparty
        session.findById("wnd[0]/usr/txtVBCOM-BSTKD").text = po_num #PO number
        session.findById("wnd[0]/usr/ctxtVBCOM-AUDAT").text = po_date #PO document date (start)
        session.findById("wnd[0]/usr/ctxtVBCOM-AUDAT_BIS").text = po_date #PO document date (end)
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[1]/usr/ctxtVBCOM-VKORG").text = salesorg #salesorg : ivy 1100 / red 1400 / AST 1300
        # session.findById("wnd[1]/usr/ctxtVBCOM-VKORG").caretPosition = 4
        session.findById("wnd[1]").sendVKey (0)
        session.findById("wnd[0]/tbar[1]/btn[32]").press()

        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER2_LAYO/shellcont/shell").currentCellRow = 6
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER2_LAYO/shellcont/shell").selectAll()
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_FL_SING").press()

        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 13
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = 12
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "13"
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 31
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = 24
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "31"
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 86
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = 78
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "86"
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
        
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 37
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = 30
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "37"
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 48
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = 39
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "48"
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 30
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "30"
        # session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()

        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 37
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = 30
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "37"
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 49
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = 42
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "49"
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = 30
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").firstVisibleRow = 21
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "30"
        session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").doubleClickCurrentCell()


        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").contextMenu()
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").selectContextMenuItem ("&XXL")
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
#            session.findById("wnd[1]/usr/ctxtDY_PATH").text = "Y:\\OM ONLY_Shared Documents\\3 Teams\\2 DP\\38. Stockcheck" #Change Location - (type 1)
        # session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"C:\Users\KISS Admin\Downloads" #Change Location - (type 1)
        # session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "input.XLSX"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 7
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        print("SAP good")
    
    except:
        print(sys.exc_info()[0])

    finally:
            session = None
            connection = None
            application = None
            SapGuiAuto = None

# %%

# Get order information from smartsheet
smartsheet_client = smartsheet.Smartsheet('rjjjwNgTfxwAjE5R5YcSKu5OocAMyLAUJa2av')
smartsheet_client.Sheets.get_sheet_as_csv(
  3482775064995716,           # sheet_id
  desktoploc+'\stock check practice')

order_df = pd.read_csv(desktoploc+'\stock check practice\Stock Check Request.csv')
order_df['Requester'] = order_df['Requester'].astype(str)

# %%
# downloadLoc = r"C:\Users\KISS Admin\Downloads"
# requestFile = downloadLoc +r"\Stock Check Request.xlsx"

# #order_df = pd.read_csv(r'C:\Users\KISS Admin\Desktop\IVYENT_DH\8. stock check automation code\Stock Check Request.csv')
# order_df = pd.read_excel(requestFile)
# order_df['Requester'] = order_df['Requester'].astype(str)

order_df['salesorg'] = order_df.apply(salesorg, axis=1)
# the first uncompleted IVY order
condition = ((order_df['Company (IVY/RED)'] == 'IVY') | (order_df['Company (IVY/RED)'] == 'IVY & RED')) & (order_df['Completed'] != True)
order_df = order_df[condition]
order_df = order_df.reset_index()

pskill()

# %%
while True:
    try:
        num=input("Enter an integer: ")
        if num=="":
            num=0
        else:
            num=int(num)
        break
    except ValueError:
        print("Invalid input. Please enter an integer.")
print(num)
if len(order_df) > 0:
    order_df = order_df.iloc[num] # TODO : change index to choose order
    # sold_to_party = order_df['Account # (Sold-to Party)']
    # po_num = order_df['PO #']
    # po_start = order_df['PO Start Date']
    # po_end = order_df['PO End Date']
    salesorg = order_df['salesorg']
    order_number=str(order_df["Order #"]).split('.')[0]

    if salesorg == '1300': #For AST orders, we do not check for plant 1000
        plant_list = ['1100', '1110']
    
    else:
        plant_list = ['1000', '1100', '1110']
    # Data Extraction From SAP

    shell = win32com.client.Dispatch("WScript.Shell")

    time.sleep(2)
    # Log in SAP QA for now / will be changed to R3 with system=P01
    user_char="-user="+data['sap_id']
    pw_char= "-pw="+data['sap_password']
    subprocess.check_call(['C:\Program Files (x86)\SAP\FrontEnd\SAPgui\\sapshcut.exe', '-system=P01', '-client=100', user_char, pw_char , 
     '-command=ZPPRMRP01', 
    '-type=Transaction', '-max'])

    # Wait until win32com detects SAPGUI
    time.sleep(5)

    # Run vbs code
    
    if __name__ == "__main__":
        sold_to_party, po_num, po_date=getfromVA03(order_number)
        main(sold_to_party, po_num, po_date,salesorg)

    '''pykill.py - selective process kill prog'''
    pskill()

    # %% Import Input file
    import os
    # print(order_number)
    # old_name=r"C:\Users\dokim2\OneDrive - Kiss Products Inc\Documents\SAP\SAP GUI\export.XLSX"

    old_name=r"C:\Users\dokim2\Documents\SAP\SAP GUI\export.XLSX"

    input_df = pd.read_excel(old_name, sheet_name='Sheet1') #Change Location - (type 1)
    # input_df = input_df[['Material', 'Order quantity', 'Plant']]
    input_df.columns = ['material', 'qty', 'plant']
    input_df_sum=input_df.groupby(['material', 'plant'])['qty'].sum().reset_index()
    input_df_sum=input_df_sum[['material', 'qty', 'plant']]
    
    input2=input_df.merge(input_df_sum, how='left', on=['material','plant'])
    input2=input2[input2.loc[:,["material","plant"]].duplicated()!=True]
    input2=input2[["material","qty_y","plant"]]
    input2.columns = ['material', 'qty', 'plant']
    input_df=input2

    input_df['plant'] = input_df['plant'].astype(str)
    input_df['material'] = input_df['material'].astype(str)

    input_df.insert(3,"order_number",order_number)
    input_df.insert(4,"salesorg",salesorg)
    input_df=input_df.loc[input_df["qty"]>0]
    # print(final_df)
    # %%
    # Export output
    # resultLoc=desktoploc+"\stock check practice"
    input_df.to_csv(r'C:\Users\dokim2\Documents\Stock Check Result'+"\\"+"simulator_input.csv",index=False) #Change Location - (type 2)
    print(order_number)

os.remove(old_name)
# %%
print("==========================================================")
print("get_stockchecklist_IVY.py done")

